"""Input loading and row-level validation for D cohorts."""

from __future__ import annotations

import csv
import gzip
import math
from pathlib import Path
from typing import Any, Mapping

from meld_icb.frozen_state_transfer import AXIS_ORDER, VALID_REPRESENTATIONS


class InputHold(ValueError):
    """Raised when source rows cannot establish a valid analysis population."""


def _read_table(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        if not sample.strip():
            return []
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters="\t,")
        except csv.Error:
            dialect = csv.excel_tab
        return [dict(row) for row in csv.DictReader(handle, dialect=dialect)]


def _parse_soft(path: Path) -> list[dict[str, str]]:
    opener = gzip.open if path.suffix.lower() == ".gz" else open
    records: list[dict[str, str]] = []
    current: dict[str, str] | None = None
    with opener(path, "rt", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            line = line.rstrip("\n")
            if line.startswith("^SAMPLE = "):
                if current is not None:
                    records.append(current)
                current = {"geo_accession": line.split(" = ", 1)[1]}
            elif current is not None and line.startswith("!Sample_title = "):
                current["sample_title"] = line.split(" = ", 1)[1]
            elif current is not None and line.startswith("!Sample_source_name_ch1 = "):
                current["source_name"] = line.split(" = ", 1)[1]
            elif current is not None and line.startswith("!Sample_characteristics_ch1 = "):
                item = line.split(" = ", 1)[1]
                if ": " in item:
                    key, value = item.split(": ", 1)
                    current[key] = value
    if current is not None:
        records.append(current)
    for index, record in enumerate(records, start=1):
        record["soft_order_position"] = str(index)
    return records


def _resolve_auxiliary(path: str, *, input_root: Path, data_root: Path | None) -> Path:
    candidate = Path(path)
    if candidate.is_absolute():
        return candidate
    if data_root is not None:
        return data_root / candidate
    return input_root / candidate


def _gse78220_rows(expression_path: Path, metadata_path: Path, contract: Mapping[str, Any]) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise InputHold("GSE78220 workbook adapter requires openpyxl") from exc
    metadata = _parse_soft(metadata_path)
    dataset_contract = contract["dataset_contract"]
    sheet_name = str(dataset_contract.get("workbook_sheet", "")).strip()
    gene_key = str(dataset_contract.get("expression_column_key", "Gene")).strip()
    sample_key = str(dataset_contract.get("metadata_sample_key", "sample_title")).strip()
    patient_key = str(dataset_contract.get("metadata_patient_key", "patient id")).strip()
    timepoint_key = str(dataset_contract.get("metadata_timepoint_key", "biopsy time")).strip()
    response_key = str(dataset_contract.get("metadata_response_key", "anti-pd-1 response")).strip()
    transform_contract = str(dataset_contract.get("transform_contract", "")).strip()
    if transform_contract != "truncate_negative_then_log2p1_before_PRE_MEAN_THEN_AXIS_Z":
        raise InputHold("workbook transform contract is missing or unsupported")
    lookup: dict[str, dict[str, str]] = {}
    for item in metadata:
        title = str(item.get(sample_key, "")).strip()
        if not title or title in lookup:
            raise InputHold(f"ambiguous GSE78220 metadata mapping: {title}")
        lookup[title] = item
    workbook = openpyxl.load_workbook(expression_path, read_only=True, data_only=True)
    if sheet_name and sheet_name not in workbook.sheetnames:
        raise InputHold(f"workbook sheet is missing: {sheet_name}")
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [str(item).strip() for item in next(rows)]
    except StopIteration as exc:
        raise InputHold("GSE78220 workbook is empty") from exc
    if not header or header[0] != gene_key:
        raise InputHold(f"workbook requires declared expression column: {gene_key}")
    columns = header[1:]
    expression: dict[str, dict[str, float]] = {column: {} for column in columns if column}
    for values in rows:
        if not values:
            continue
        gene = str(values[0]).strip().upper()
        if not gene:
            continue
        for column, value in zip(columns, values[1:]):
            if not column or value in (None, ""):
                continue
            try:
                number = max(0.0, float(value))
            except (TypeError, ValueError) as exc:
                raise InputHold(f"non-numeric GSE78220 value for {gene}/{column}") from exc
            expression[column][gene] = number
    response = contract["response"]
    accepted = {str(item).strip().upper() for item in response["positive_labels"] + response["negative_labels"]}
    rows_out: list[dict[str, Any]] = []
    for column in columns:
        title = column.split(".", 1)[0]
        record = lookup.get(title)
        if record is None:
            raise InputHold(f"GSE78220 expression column has no metadata: {column}")
        response_label = str(record.get(response_key, "")).strip()
        timepoint = str(record.get(timepoint_key, "")).strip().lower()
        if timepoint != "pre-treatment" or response_label.upper() not in accepted:
            continue
        patient = str(record.get(patient_key, "")).strip()
        if not patient:
            raise InputHold(f"GSE78220 row has no patient mapping: {column}")
        rows_out.append({
            "sample_id": column,
            "patient_id": patient,
            "response": response_label,
            "timepoint": timepoint,
            "geo_accession": record.get("geo_accession", ""),
            "_expression": expression[column],
        })
    if not rows_out:
        raise InputHold("GSE78220 has no eligible pretreatment rows")
    return rows_out


def _generic_rows(path: Path) -> list[dict[str, Any]]:
    rows = _read_table(path)
    for row in rows:
        expression: dict[str, float] = {}
        for key, value in row.items():
            if key in {"sample_id", "patient_id", "response", "timepoint", "normalization_scope_id", "normalization_source", "analysis_unit", "cohort", "pair_id"}:
                continue
            if str(value).strip() == "":
                continue
            try:
                expression[key] = float(value)
            except (TypeError, ValueError):
                continue
        row["_expression"] = expression
    return rows


def load_rows(*, input_path: Path, contract: Mapping[str, Any], data_root: Path | None = None) -> list[dict[str, Any]]:
    adapter = str(contract["input_adapter"])
    if adapter == "gse78220_workbook_soft":
        metadata_files = contract.get("metadata_files", [])
        if not metadata_files:
            raise InputHold("GSE78220 adapter requires metadata_files")
        metadata = _resolve_auxiliary(str(metadata_files[0]), input_root=input_path.parent, data_root=data_root)
        if not metadata.exists():
            raise InputHold(f"GSE78220 metadata input is missing: {metadata}")
        return _gse78220_rows(input_path, metadata, contract)
    if adapter in {"sade_feldman_candidate", "imvigor210_candidate", "gse282471_superseded"}:
        raise InputHold(f"{adapter} source adapter is not authority-closed")
    if adapter != "tabular_rows":
        raise InputHold(f"unsupported source adapter: {adapter}")
    return _generic_rows(input_path)


def validate_rows(rows: list[Mapping[str, Any]], contract: Mapping[str, Any], *, representation: str) -> None:
    if representation not in VALID_REPRESENTATIONS:
        raise InputHold(f"representation must be explicit: {VALID_REPRESENTATIONS}")
    if not rows:
        raise InputHold("cohort input is empty")
    response_column = str(contract["response"]["column"])
    seen_samples: set[str] = set()
    for row in rows:
        missing = [field for field in ("sample_id", "patient_id", response_column) if not str(row.get(field, "")).strip()]
        if missing:
            raise InputHold(f"missing required row fields: {missing}")
        sample_id = str(row["sample_id"])
        if sample_id in seen_samples:
            raise InputHold(f"duplicate sample_id: {sample_id}")
        seen_samples.add(sample_id)
        value = str(row[response_column]).strip().upper()
        labels = {str(item).strip().upper() for item in contract["response"]["positive_labels"] + contract["response"]["negative_labels"]}
        if value not in labels:
            raise InputHold(f"unsupported response label: {row[response_column]}")
        expression = row.get("_expression", {})
        if representation == "gene_expression" and not expression:
            raise InputHold(f"sample {sample_id} has no numeric expression")
        if representation == "canonical_axis_z":
            absent = [axis for axis in AXIS_ORDER if not str(row.get(axis, "")).strip()]
            provenance = [field for field in ("normalization_scope_id", "normalization_source") if not str(row.get(field, "")).strip()]
            if absent or provenance:
                raise InputHold(f"canonical_axis_z provenance/axes missing: {absent + provenance}")


def encode_response(value: Any, contract: Mapping[str, Any]) -> int:
    normalized = str(value).strip().upper()
    positives = {str(item).strip().upper() for item in contract["response"]["positive_labels"]}
    negatives = {str(item).strip().upper() for item in contract["response"]["negative_labels"]}
    if normalized in positives:
        return 1
    if normalized in negatives:
        return 0
    raise InputHold(f"unsupported response label: {value}")
